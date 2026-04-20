"""Generate consolidated XLSX — one row per parcela (mirrors the web preview table).

Supports optional column filtering via the `columns` parameter.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from parser import ExtractResult


MONEY_FMT = "#,##0.00"
PCT_FMT = "0.0000"

TITLE_FONT = Font(bold=True, size=14, color="1E3A8A")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="334155")
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="E0E7FF")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN = Side(style="thin", color="CBD5E1")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TOP_BORDER = Border(top=Side(style="medium", color="334155"))

# (key, header_label, width)
ALL_COLUMNS = [
    ("num", "#", 6),
    ("grupo", "Grupo", 10),
    ("cota", "Cota", 12),
    ("contrato", "Contrato", 16),
    ("emissao", "Emissão", 14),
    ("prazo", "Prazo", 10),
    ("parcelas", "Parcelas", 10),
    ("vencto", "Vencto", 14),
    ("pagto", "Pagto", 14),
    ("vl_credito", "Vl. Crédito", 16),
    ("vl_devido", "Vl. Devido", 14),
    ("vl_pago", "Vl. Pago", 14),
    ("pct_pago", "% Pago", 10),
    ("quota_consorcio", "Quota Consórcio", 16),
    ("fundo_reserva", "Fundo Reserva", 16),
    ("taxa_adm", "Taxa ADM", 16),
]


def _money(cell, value):
    cell.value = value
    cell.number_format = MONEY_FMT
    cell.alignment = RIGHT


def _date(cell, value: str):
    try:
        cell.value = datetime.strptime(value, "%d/%m/%Y")
        cell.number_format = "DD/MM/YYYY"
    except (ValueError, TypeError):
        cell.value = value
    cell.alignment = CENTER


def build_xlsx(
    results: Iterable[ExtractResult],
    sheet_name: str = "Extratos",
    columns: list[str] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    results = list(results)

    # Determine active columns
    if columns:
        active = [c for c in ALL_COLUMNS if c[0] in columns]
    else:
        active = list(ALL_COLUMNS)

    keys = [c[0] for c in active]

    def col_of(key: str) -> int | None:
        try:
            return keys.index(key) + 1
        except ValueError:
            return None

    # Column widths
    for i, (_, _, w) in enumerate(active, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Flatten: one row per parcela
    rows_data = []
    for r in results:
        parcelas = sorted(r.conta_corrente, key=lambda p: int(p.ass))
        for p in parcelas:
            rows_data.append((r, p))

    n = len(rows_data)
    first_data = 2
    last_data = first_data + n - 1

    # Header row
    hdr_row = 1
    for i, (_, label, _) in enumerate(active, 1):
        c = ws.cell(row=hdr_row, column=i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BOX

    # Data rows
    for idx, (r, p) in enumerate(rows_data):
        row = first_data + idx
        for ci_idx, key in enumerate(keys):
            col = ci_idx + 1
            if key == "num":
                ws.cell(row=row, column=col, value=idx + 1).alignment = CENTER
            elif key == "grupo":
                ws.cell(row=row, column=col, value=r.grupo.lstrip("0") or "0").alignment = CENTER
            elif key == "cota":
                ws.cell(row=row, column=col, value=r.cota).alignment = CENTER
            elif key == "contrato":
                ws.cell(row=row, column=col, value=r.contrato).alignment = CENTER
            elif key == "emissao":
                _date(ws.cell(row=row, column=col), r.data_emissao)
            elif key == "prazo":
                prazo = f"{str(r.qtde_parcelas_pagas).zfill(3)}/{r.prazo_total}"
                ws.cell(row=row, column=col, value=prazo).alignment = CENTER
            elif key == "parcelas":
                ws.cell(row=row, column=col, value=p.ass).alignment = CENTER
            elif key == "vencto":
                _date(ws.cell(row=row, column=col), p.vencto)
            elif key == "pagto":
                _date(ws.cell(row=row, column=col), p.pagto)
            elif key == "vl_credito":
                _money(ws.cell(row=row, column=col), p.vl_cred)
            elif key == "vl_devido":
                _money(ws.cell(row=row, column=col), p.vl_devido)
            elif key == "vl_pago":
                _money(ws.cell(row=row, column=col), p.vl_pago)
            elif key == "pct_pago":
                cell = ws.cell(row=row, column=col, value=p.pct_pago)
                cell.number_format = PCT_FMT
                cell.alignment = RIGHT
            elif key == "quota_consorcio":
                _money(ws.cell(row=row, column=col), r.valores_pagos.fundo_comum)
            elif key == "fundo_reserva":
                _money(ws.cell(row=row, column=col), r.valores_pagos.fundo_reserva)
            elif key == "taxa_adm":
                _money(ws.cell(row=row, column=col), r.valores_pagos.taxa_administracao)

    # Totals row
    sum_keys = {"vl_credito", "vl_devido", "vl_pago", "quota_consorcio", "fundo_reserva", "taxa_adm"}
    if n:
        trow = last_data + 1
        for ci_idx, key in enumerate(keys):
            col = ci_idx + 1
            cell = ws.cell(row=trow, column=col)
            if key in sum_keys:
                letter = get_column_letter(col)
                cell.value = f"=SUM({letter}{first_data}:{letter}{last_data})"
                cell.number_format = MONEY_FMT
                cell.alignment = RIGHT
            cell.fill = TOTAL_FILL
            cell.font = TOTAL_FONT
            cell.border = TOP_BORDER

        # Label "TOTAL"
        label_col = col_of("prazo") or col_of("contrato") or col_of("cota") or 1
        ws.cell(row=trow, column=label_col, value="TOTAL").font = TOTAL_FONT

    ws.freeze_panes = f"A{first_data}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
